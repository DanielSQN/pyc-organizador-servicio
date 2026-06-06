from __future__ import annotations

from copy import copy
from datetime import datetime
from io import BytesIO
from textwrap import dedent

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.loader import (
    clean_volunteers,
    get_available_volunteers,
    get_summary,
    get_unavailable_volunteers,
    read_volunteers,
    validate_volunteers,
)
from src.rules import (
    GRUPO_PRIMER_SERVICIO_DEFAULT,
    SUPERVISORES_ZONA,
    TITULOS_ZONA,
    TURNOS,
    ZONAS_CONFIGURADAS,
    build_turnos_por_grupo,
    get_positions_df,
)
from src.scheduler import FREE_NAME, UNASSIGNED_NAME, apply_manual_assignment, generate_schedule, validate_schedule


COORDINADOR_NOMBRE = "FERNANDO BLANCO G."


def inject_app_styles() -> None:
    st.markdown(
        """
        <style>
        :root {
            --pyc-blue: #1463f3;
            --pyc-blue-soft: #edf5ff;
            --pyc-green: #2fbf71;
            --pyc-ink: #172033;
            --pyc-muted: #687387;
            --pyc-border: #d9e2ef;
            --pyc-bg: #f6f8fb;
        }
        html {
            color-scheme: light !important;
        }
        body,
        .stApp,
        [data-testid="stAppViewContainer"],
        [data-testid="stMain"],
        [data-testid="stVerticalBlock"],
        [data-testid="stHorizontalBlock"] {
            background: var(--pyc-bg) !important;
            color: var(--pyc-ink) !important;
        }
        h1, h2, h3, h4, h5, h6,
        p, label,
        [data-testid="stMarkdownContainer"],
        [data-testid="stMarkdownContainer"] p,
        [data-testid="stMarkdownContainer"] span {
            color: var(--pyc-ink) !important;
        }
        header,
        header[data-testid="stHeader"],
        [data-testid="stDecoration"],
        div[data-testid="stToolbar"],
        #MainMenu,
        footer {
            display: none;
        }
        .block-container {
            padding: 0.55rem 0.7rem 1.25rem;
            max-width: none;
        }
        div[data-testid="stHorizontalBlock"] {
            align-items: stretch;
        }
        div[data-testid="stMetric"] {
            border: 1px solid var(--pyc-border);
            border-radius: 10px;
            padding: 0.85rem 1rem;
            background: #ffffff !important;
            color: var(--pyc-ink) !important;
            box-shadow: 0 8px 24px rgba(15, 23, 42, 0.04);
        }
        div[data-testid="stMetric"] label {
            color: var(--pyc-muted) !important;
        }
        .pyc-panel {
            border: 1px solid var(--pyc-border);
            border-radius: 14px;
            padding: 1.15rem;
            background: #ffffff !important;
            color: var(--pyc-ink) !important;
            box-shadow: 0 12px 32px rgba(15, 23, 42, 0.05);
        }
        .pyc-nav {
            background: #ffffff !important;
            color: var(--pyc-ink) !important;
            border: 1px solid var(--pyc-border);
            border-radius: 16px;
            padding: 0.9rem 1rem 0.75rem;
            box-shadow: 0 12px 32px rgba(15, 23, 42, 0.05);
            margin-bottom: 0.9rem;
        }
        .pyc-brand {
            display: flex;
            align-items: center;
            gap: 0.75rem;
        }
        .pyc-brand-icon {
            width: 42px;
            height: 42px;
            display: grid;
            place-items: center;
            border-radius: 10px;
            color: var(--pyc-blue);
            background: var(--pyc-blue-soft);
            font-size: 1.35rem;
            font-weight: 800;
        }
        .pyc-user {
            display: flex;
            align-items: center;
            justify-content: flex-end;
            gap: 0.65rem;
            color: var(--pyc-ink);
            min-height: 54px;
        }
        .pyc-avatar {
            width: 38px;
            height: 38px;
            border-radius: 50%;
            background: #dbeafe;
            color: var(--pyc-blue);
            display: grid;
            place-items: center;
            font-weight: 750;
        }
        .pyc-nav-row {
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 1rem;
            margin-bottom: 0.9rem;
        }
        .pyc-horizontal-steps {
            margin-bottom: 1rem;
        }
        .pyc-step-arrow {
            color: #9ca3af !important;
            font-size: 1.3rem;
            font-weight: 900;
            text-align: center;
            padding-top: 0.85rem;
        }
        .pyc-file-uploader div[data-testid="stFileUploader"] {
            max-width: 430px;
        }
        div[data-testid="stFileUploader"] {
            max-width: 430px;
        }
        .pyc-file-uploader section[data-testid="stFileUploaderDropzone"] {
            padding: 0.55rem !important;
            min-height: 4.5rem !important;
        }
        section[data-testid="stFileUploaderDropzone"] {
            padding: 0.55rem !important;
            min-height: 4.5rem !important;
        }
        .pyc-file-uploader section[data-testid="stFileUploaderDropzone"] button {
            min-height: 2.15rem !important;
            border-radius: 9px !important;
            font-size: 0.86rem !important;
        }
        section[data-testid="stFileUploaderDropzone"] button {
            min-height: 2.15rem !important;
            border-radius: 9px !important;
            font-size: 0.86rem !important;
        }
        div[data-testid="stButton"] > button,
        div[data-testid="stDownloadButton"] > button,
        button[data-testid="stBaseButton-secondary"],
        button[data-testid="stBaseButton-primary"] {
            border-radius: 14px;
            min-height: 3.25rem;
            font-weight: 800;
            background: #ffffff !important;
            color: var(--pyc-ink) !important;
            border: 1px solid var(--pyc-border) !important;
        }
        div[data-testid="stButton"] > button {
            box-shadow: 0 8px 20px rgba(15, 23, 42, 0.04);
        }
        div[data-testid="stButton"] > button[kind="primary"],
        button[data-testid="stBaseButton-primary"] {
            background: var(--pyc-blue-soft) !important;
            color: var(--pyc-blue) !important;
            border: 1px solid #b8d7ff !important;
            box-shadow: 0 8px 20px rgba(15, 23, 42, 0.04);
        }
        div[data-testid="stButton"] > button[kind="secondary"]:not(:disabled) {
            background: #f0fdf4 !important;
            color: #166534 !important;
            border: 1px solid #86efac !important;
        }
        div[data-testid="stButton"] > button:disabled {
            background: #f8fafc !important;
            color: #94a3b8 !important;
            border: 1px solid #e5e7eb !important;
            opacity: 1 !important;
            box-shadow: none;
        }
        div[data-testid="stButton"] > button p,
        div[data-testid="stButton"] > button span {
            color: inherit !important;
        }
        div[data-testid="stDownloadButton"] > button {
            min-height: 2.5rem !important;
            border-radius: 10px !important;
            font-size: 0.9rem !important;
        }
        button[data-baseweb="tab"] {
            color: var(--pyc-ink) !important;
        }
        button[data-baseweb="tab"][aria-selected="true"] {
            color: #ff4b4b !important;
        }
        .pyc-title {
            font-size: 1.35rem;
            font-weight: 750;
            margin-bottom: 0.1rem;
            color: var(--pyc-ink);
        }
        .pyc-subtitle {
            color: var(--pyc-muted) !important;
            margin-bottom: 0;
        }
        .pyc-sidebar {
            min-height: 620px;
        }
        .pyc-steps {
            position: relative;
            padding: 0.4rem 0 0.4rem 0.15rem;
        }
        .pyc-steps::before {
            content: "";
            position: absolute;
            left: 1.18rem;
            top: 1.6rem;
            bottom: 2rem;
            width: 2px;
            background: #d6dde8;
        }
        .pyc-step {
            display: grid;
            grid-template-columns: 2.3rem 1fr;
            gap: 0.55rem;
            position: relative;
            padding: 0.7rem 0.7rem 0.7rem 0;
            color: var(--pyc-ink);
            font-size: 0.95rem;
            border-radius: 10px;
            margin-bottom: 0.35rem;
        }
        .pyc-step-active {
            background: var(--pyc-blue-soft);
            color: var(--pyc-blue);
            font-weight: 750;
        }
        .pyc-step-marker {
            width: 1.95rem;
            height: 1.95rem;
            border-radius: 999px;
            display: grid;
            place-items: center;
            font-weight: 800;
            z-index: 1;
            border: 1px solid #cfd8e6;
            background: #ffffff !important;
            color: #7a8597 !important;
        }
        .pyc-step-done .pyc-step-marker {
            background: var(--pyc-green) !important;
            border-color: var(--pyc-green) !important;
            color: #ffffff !important;
        }
        .pyc-step-active .pyc-step-marker {
            background: var(--pyc-blue) !important;
            border-color: var(--pyc-blue) !important;
            color: #ffffff !important;
        }
        .pyc-step-title {
            font-weight: 750;
            margin-top: 0.12rem;
        }
        .pyc-step small {
            color: var(--pyc-muted) !important;
            display: block;
            margin-top: 0.25rem;
            font-weight: 500;
        }
        .pyc-help-card {
            margin-top: 4rem;
            border: 1px solid #bfdbfe;
            border-radius: 10px;
            padding: 1rem;
            background: #f8fbff;
            color: var(--pyc-ink);
        }
        .pyc-help-card strong {
            color: var(--pyc-blue);
            display: block;
            margin-bottom: 0.6rem;
        }
        .pyc-help-card p {
            color: #4b5563;
            font-size: 0.9rem;
            line-height: 1.45;
            margin-bottom: 0.8rem;
        }
        .pyc-help-card a {
            color: var(--pyc-blue);
            text-decoration: none;
            font-weight: 650;
        }
        .pyc-alert-card {
            border: 1px solid #e5e7eb;
            border-radius: 6px;
            padding: 0.65rem 0.75rem;
            margin-bottom: 0.6rem;
            background: #ffffff !important;
            color: var(--pyc-ink) !important;
            font-size: 0.9rem;
        }
        .pyc-alert-danger {
            border-left: 4px solid #dc2626;
        }
        .pyc-alert-warning {
            border-left: 4px solid #d97706;
        }
        .pyc-alert-info {
            border-left: 4px solid #2563eb;
        }
        .pyc-volunteer-strip {
            display: flex;
            align-items: center;
            gap: 0.65rem;
            flex-wrap: wrap;
            margin: 0.4rem 0 0.65rem;
        }
        .pyc-volunteer-chip {
            border: 1px solid var(--pyc-border);
            border-radius: 999px;
            background: #ffffff !important;
            color: var(--pyc-muted) !important;
            padding: 0.35rem 0.65rem;
            font-size: 0.86rem;
            line-height: 1;
        }
        .pyc-volunteer-chip strong {
            color: var(--pyc-ink) !important;
            margin-left: 0.25rem;
        }
        .pyc-volunteer-chip.pyc-group-chip-a {
            background: #2563eb !important;
            border-color: #2563eb !important;
            color: #ffffff !important;
        }
        .pyc-volunteer-chip.pyc-group-chip-b {
            background: #16a34a !important;
            border-color: #16a34a !important;
            color: #ffffff !important;
        }
        .pyc-volunteer-chip.pyc-group-chip-a strong,
        .pyc-volunteer-chip.pyc-group-chip-b strong {
            color: #ffffff !important;
        }
        .pyc-file-chip-row {
            display: flex;
            align-items: center;
            gap: 0.5rem;
            flex-wrap: wrap;
            margin: 0.25rem 0 0.9rem;
        }
        .pyc-file-chip {
            border: 1px solid var(--pyc-border);
            border-radius: 999px;
            background: #ffffff !important;
            color: var(--pyc-muted) !important;
            padding: 0.42rem 0.75rem;
            font-size: 0.88rem;
            line-height: 1;
        }
        .pyc-file-chip strong {
            color: var(--pyc-ink) !important;
        }
        .pyc-status-ok {
            border-color: #bbf7d0 !important;
            background: #f0fdf4 !important;
            color: #166534 !important;
        }
        .pyc-status-ok strong {
            color: #166534 !important;
        }
        .pyc-status-warn {
            border-color: #fde68a !important;
            background: #fffbeb !important;
            color: #92400e !important;
        }
        .pyc-status-warn strong {
            color: #92400e !important;
        }
        .pyc-status-error {
            border-color: #fecaca !important;
            background: #fef2f2 !important;
            color: #991b1b !important;
        }
        .pyc-status-error strong {
            color: #991b1b !important;
        }
        .pyc-download-col div[data-testid="stDownloadButton"] > button {
            min-height: 2.5rem;
            border-radius: 10px;
            padding-left: 0.8rem;
            padding-right: 0.8rem;
        }
        div[data-testid="stDownloadButton"] > button {
            background: #1463f3 !important;
            color: #ffffff !important;
            border-color: #1463f3 !important;
            box-shadow: 0 10px 22px rgba(20, 99, 243, 0.18);
        }
        div[data-testid="stDownloadButton"] > button p,
        div[data-testid="stDownloadButton"] > button span {
            color: #ffffff !important;
        }
        div[data-testid="stFormSubmitButton"] > button {
            min-height: 2.65rem !important;
            border-radius: 10px !important;
            background: #1463f3 !important;
            color: #ffffff !important;
            border: 1px solid #1463f3 !important;
            box-shadow: 0 10px 22px rgba(20, 99, 243, 0.18);
            font-weight: 800;
        }
        div[data-testid="stFormSubmitButton"] > button p,
        div[data-testid="stFormSubmitButton"] > button span {
            color: #ffffff !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def build_output_excel(
    schedule_df: pd.DataFrame,
    alerts: list[str],
    available_df: pd.DataFrame,
    supervisor_config: dict[str, dict[str, str]] | None = None,
) -> bytes:
    """Crea el Excel final en memoria con una vista operativa por zonas."""
    output = BytesIO()
    alerts_df = pd.DataFrame({"Alerta": alerts})

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_formatted_schedule_sheet(writer, schedule_df, supervisor_config or SUPERVISORES_ZONA)
        alerts_df.to_excel(writer, sheet_name="Alertas", index=False)
        available_df.to_excel(writer, sheet_name="Voluntarios Disponibles", index=False)
        apply_report_styles(writer.sheets["Programacion"])

    output.seek(0)
    return output.getvalue()


def _write_formatted_schedule_sheet(
    writer: pd.ExcelWriter,
    schedule_df: pd.DataFrame,
    supervisor_config: dict[str, dict[str, str]],
) -> None:
    workbook = writer.book
    worksheet = workbook.create_sheet("Programacion")
    writer.sheets["Programacion"] = worksheet

    turnos = sorted(TURNOS)
    last_col = 2 + len(turnos)
    metadata_col = last_col + 1

    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col)
    worksheet.cell(row=1, column=2, value="PROGRAMACIÓN OPERATIVA PYC")

    worksheet.cell(row=2, column=2, value=f"Coordinador:  {COORDINADOR_NOMBRE}")

    current_row = 4
    for zone in _ordered_zones(schedule_df):
        zone_df = schedule_df[schedule_df["Zona"] == zone]
        if zone_df.empty:
            continue

        matrix_df = build_schedule_matrix(zone_df).drop(columns=["Zona"])
        supervisor_info = supervisor_config.get(zone, {})
        supervisor = supervisor_info.get("supervisor", "")
        asistente = supervisor_info.get("asistente", "")

        worksheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=last_col)
        worksheet.cell(row=current_row, column=2, value=TITULOS_ZONA.get(zone, f"ZONA {zone}"))
        current_row += 1

        worksheet.cell(row=current_row, column=2, value=f"Supervisor: {supervisor}")
        worksheet.cell(row=current_row, column=4, value=f"Asistente: {asistente}" if asistente else "Asistente:")
        current_row += 1

        headers = ["POSICIÓN", *[_turn_header(turno) for turno in turnos]]
        for col_index, header in enumerate(headers, start=2):
            worksheet.cell(row=current_row, column=col_index, value=header)
        current_row += 1

        for _, row in matrix_df.iterrows():
            position_cell = worksheet.cell(row=current_row, column=2, value=row["Posición"])
            worksheet.cell(row=current_row, column=metadata_col, value=row["Tipo"])

            for col_offset, turno in enumerate(turnos, start=3):
                value = row[_turn_key(turno)]
                worksheet.cell(row=current_row, column=col_offset, value=value)

            current_row += 1

        current_row += 3


def apply_report_styles(ws) -> None:
    """Aplica estilos de reporte operativo a la hoja de programación."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B6"
    metadata_col = ws.max_column if _has_type_metadata(ws) else None
    last_col = metadata_col - 1 if metadata_col else ws.max_column
    if metadata_col:
        ws.column_dimensions[get_column_letter(metadata_col)].hidden = True

    fills = {
        "title": PatternFill("solid", fgColor="17365D"),
        "coordinator": PatternFill("solid", fgColor="E5E7EB"),
        "header": PatternFill("solid", fgColor="404040"),
        "refuerzo": PatternFill("solid", fgColor="DDEBFF"),
        "free": PatternFill("solid", fgColor="F3F4F6"),
        "unassigned": PatternFill("solid", fgColor="FCE4E4"),
        "group_a": PatternFill("solid", fgColor="D9EAF7"),
        "group_b": PatternFill("solid", fgColor="E2F0D9"),
        "white": PatternFill("solid", fgColor="FFFFFF"),
    }
    soft_side = Side(style="thin", color="D9E2EF")
    soft_border = Border(left=soft_side, right=soft_side, top=soft_side, bottom=soft_side)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 36
    for col_idx in range(3, last_col + 1):
        column = get_column_letter(col_idx)
        ws.column_dimensions[column].width = 30

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=last_col):
        for cell in row:
            cell.border = soft_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.fill.fill_type is None:
                cell.fill = fills["white"]
        ws.row_dimensions[row[0].row].height = 24

    ws.row_dimensions[1].height = 30

    title_cell = ws["B1"]
    title_cell.fill = fills["title"]
    title_cell.font = Font(color="FFFFFF", bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(2, last_col + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = fills["coordinator"]
        cell.font = Font(bold=True, size=12, color="111827")

    for row_idx in range(4, ws.max_row + 1):
        row_label = str(ws.cell(row=row_idx, column=2).value or "")

        if row_label.startswith("ZONA"):
            _style_full_report_row(
                ws,
                row_idx,
                _zone_fill(row_label),
                Font(color="FFFFFF", bold=True, size=14),
                last_col=last_col,
            )
            ws.row_dimensions[row_idx].height = 26
            continue

        if row_label.startswith("Supervisor:"):
            _style_full_report_row(
                ws,
                row_idx,
                fills["coordinator"],
                Font(bold=True, size=12, color="111827"),
                last_col=last_col,
            )
            continue

        if row_label == "POSICIÓN":
            _style_full_report_row(
                ws,
                row_idx,
                fills["header"],
                Font(color="FFFFFF", bold=True, size=12),
                horizontal="center",
                last_col=last_col,
            )
            ws.row_dimensions[row_idx].height = 24
            continue

        if not row_label:
            continue

        position_cell = ws.cell(row=row_idx, column=2)
        row_type = str(ws.cell(row=row_idx, column=metadata_col).value or "") if metadata_col else ""
        if row_type == "refuerzo" or "refuerzo" in row_label.lower() or "(ref)" in row_label.lower():
            position_cell.fill = fills["refuerzo"]
            position_cell.font = Font(color="2563EB", bold=True, size=12)
        else:
            position_cell.font = Font(color="111827", bold=True, size=12)

        for col_idx in range(3, last_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = str(cell.value or "")
            if not value:
                continue
            if FREE_NAME in value:
                cell.fill = fills["free"]
                cell.font = Font(color="6B7280", italic=True, size=12)
            elif UNASSIGNED_NAME in value:
                cell.fill = fills["unassigned"]
                cell.font = Font(color="B91C1C", bold=True, size=12)
            elif "(A" in value:
                cell.fill = fills["group_a"]
                cell.font = Font(color="1F3B63", size=12)
            elif "(B" in value:
                cell.fill = fills["group_b"]
                cell.font = Font(color="31572C", size=12)


def _style_full_report_row(
    ws,
    row_idx: int,
    fill: PatternFill,
    font: Font,
    horizontal: str = "left",
    last_col: int | None = None,
) -> None:
    last_col = last_col or ws.max_column
    for col_idx in range(2, last_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.font = copy(font)
        cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)


def _has_type_metadata(ws) -> bool:
    values = {
        str(ws.cell(row=row_idx, column=ws.max_column).value or "")
        for row_idx in range(1, ws.max_row + 1)
    }
    return bool(values & {"critica", "refuerzo"})


def _zone_fill(row_label: str) -> PatternFill:
    palette = {
        "Z1": "1F4E78",
        "Z2": "375623",
        "Z3": "5B2C6F",
    }
    zone = _zone_from_title(row_label)
    color = palette.get(zone, "374151")
    return PatternFill("solid", fgColor=color)


def _zone_from_title(row_label: str) -> str:
    for zone in ZONAS_CONFIGURADAS:
        if f"({zone})" in row_label or row_label.endswith(zone) or f" {zone} " in row_label:
            return zone
    return row_label.split()[-1] if row_label.split() else ""


def show_summary(summary: dict) -> None:
    first_row = st.columns(3)
    first_row[0].metric("Total", summary["total_cargados"])
    first_row[1].metric("Disponibles", summary["total_disponibles"])
    first_row[2].metric("No disponibles", summary["total_no_disponibles"])

    group_counts = summary.get("grupos_disponibles", {})
    second_row = st.columns(max(len(group_counts) + 2, 2))
    for index, (grupo, total) in enumerate(group_counts.items()):
        second_row[index].metric(f"Grupo {grupo}", total)
    second_row[len(group_counts)].metric("Hombres", summary["hombres_disponibles"])
    second_row[len(group_counts) + 1].metric("Mujeres", summary["mujeres_disponibles"])


def show_summary_chips(summary: dict) -> None:
    group_chips = "".join(
        f'<span class="pyc-volunteer-chip">Grupo {grupo} <strong>{total}</strong></span>'
        for grupo, total in summary.get("grupos_disponibles", {}).items()
    )
    st.markdown(
        f"""
        <div class="pyc-volunteer-strip">
          <span class="pyc-volunteer-chip">Total <strong>{summary["total_cargados"]}</strong></span>
          <span class="pyc-volunteer-chip">Disponibles <strong>{summary["total_disponibles"]}</strong></span>
          <span class="pyc-volunteer-chip">No disponibles <strong>{summary["total_no_disponibles"]}</strong></span>
          {group_chips}
          <span class="pyc-volunteer-chip">Hombres <strong>{summary["hombres_disponibles"]}</strong></span>
          <span class="pyc-volunteer-chip">Mujeres <strong>{summary["mujeres_disponibles"]}</strong></span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def show_brand() -> None:
    st.markdown(
        """
        <div class="pyc-brand">
          <div class="pyc-brand-icon">✓</div>
          <div>
            <div class="pyc-title">Organizador de Servicio</div>
            <p class="pyc-subtitle">Asignación automática</p>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def show_user_badge() -> None:
    st.markdown(
        """
        <div class="pyc-user">
          <div class="pyc-avatar">U</div>
          <div>
            <strong>Coordinador</strong>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def show_top_nav(current_step: int, max_step: int, has_step_error: bool = False) -> None:
    steps = [
        "Cargar y validar",
        "Generar programación",
        "Revisar y descargar",
    ]

    st.markdown(
        dedent(
            f"""
            <div class="pyc-nav">
              <div class="pyc-nav-row">
                <div class="pyc-brand">
                  <div class="pyc-brand-icon">✓</div>
                  <div>
                    <div class="pyc-title">Organizador de Servicio</div>
                    <p class="pyc-subtitle">Asignación automática</p>
                  </div>
                </div>
                <div class="pyc-user">
                  <div class="pyc-avatar">U</div>
                  <div><strong>Coordinador</strong></div>
                </div>
              </div>
            </div>
            """
        ).strip(),
        unsafe_allow_html=True,
    )

    if has_step_error:
        st.markdown(
            """
            <style>
            div[data-testid="stButton"] > button[kind="primary"] {
                background: #fef2f2 !important;
                color: #b91c1c !important;
                border: 1px solid #fca5a5 !important;
            }
            div[data-testid="stButton"] > button[kind="primary"] p,
            div[data-testid="stButton"] > button[kind="primary"] span {
                color: #b91c1c !important;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )

    st.markdown('<div class="pyc-horizontal-steps">', unsafe_allow_html=True)
    step_cols = st.columns([1, 0.04, 1, 0.04, 1], gap="small")
    step_positions = [0, 2, 4]

    for index, (title, column_index) in enumerate(zip(steps, step_positions), start=1):
        completed = index < current_step
        disabled = index > max_step
        marker = "✓" if completed else "●" if index == current_step else "○"
        label = f"{marker}  {index}. {title}"
        button_type = "primary" if index == current_step else "secondary"

        with step_cols[column_index]:
            if st.button(label, key=f"nav_step_{index}", disabled=disabled, type=button_type, width="stretch"):
                st.query_params["step"] = str(index)
                st.rerun()

        if index < len(steps):
            with step_cols[column_index + 1]:
                st.markdown('<div class="pyc-step-arrow">→</div>', unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)


def show_loaded_file_badge(uploaded_file) -> None:
    st.markdown(
        f"""
        <div class="pyc-panel" style="padding:0.72rem 0.9rem; min-height:54px;">
          <strong>Archivo cargado</strong><br>
          <small>{uploaded_file.name}</small>
        </div>
        """,
        unsafe_allow_html=True,
    )


def clear_uploaded_file() -> None:
    for key in [
        "uploaded_file_bytes",
        "uploaded_file_name",
        "uploaded_sheet_name",
        "schedule_df",
        "alerts",
        "available_df",
        "active_step",
    ]:
        st.session_state.pop(key, None)
    st.query_params.clear()


def handle_uploaded_file(uploaded_file, target_step: int = 2) -> None:
    st.session_state["uploaded_file_bytes"] = uploaded_file.getvalue()
    st.session_state["uploaded_file_name"] = uploaded_file.name
    st.session_state.pop("schedule_df", None)
    st.session_state.pop("alerts", None)
    st.session_state.pop("available_df", None)
    st.query_params["step"] = str(target_step)
    st.rerun()


def get_requested_step(max_step: int) -> int:
    raw_step = st.query_params.get("step")
    try:
        requested_step = int(raw_step) if raw_step is not None else max_step
    except ValueError:
        requested_step = max_step

    return min(max(requested_step, 1), max_step)


def show_step_panel(current_step: int) -> None:
    steps = [
        ("Cargar voluntarios", "Subir archivo Excel"),
        ("Validar datos", "Revisar estructura"),
        ("Generar programación", "Aplicar reglas"),
        ("Revisar y descargar", "Exportar resultado"),
    ]

    step_items = []
    for index, (title, detail) in enumerate(steps, start=1):
        state_class = "pyc-step-done" if index < current_step else "pyc-step-active" if index == current_step else ""
        marker = "✓" if index < current_step else str(index)
        step_items.append(
            f"""
            <div class="pyc-step {state_class}">
              <div class="pyc-step-marker">{marker}</div>
              <div>
                <div class="pyc-step-title">{index}. {title}</div>
                <small>{detail}</small>
              </div>
            </div>
            """
        )

    st.markdown(
        dedent(
            f"""
            <div class="pyc-panel pyc-sidebar">
              <div class="pyc-steps">
                {''.join(step_items)}
              </div>
              <div class="pyc-help-card">
                <strong>¿Cómo funciona?</strong>
                <p>Leemos el Excel de voluntarios, filtramos automáticamente los estados válidos
                y aplicamos las reglas de negocio para generar la programación final.</p>
                <a href="#">Ver reglas de negocio</a>
              </div>
            </div>
            """
        ).strip(),
        unsafe_allow_html=True,
    )


def show_file_summary(uploaded_file, summary: dict, sheet_name: str, status: str = "ok") -> None:
    status_meta = {
        "ok": ("Cargado correctamente", "pyc-status-ok"),
        "warn": ("Revisar alertas", "pyc-status-warn"),
        "error": ("Con errores", "pyc-status-error"),
    }
    status_text, status_class = status_meta.get(status, status_meta["ok"])

    st.markdown(
        f"""
        <div class="pyc-file-chip-row">
          <span class="pyc-file-chip">Archivo: <strong>{uploaded_file.name}</strong></span>
          <span class="pyc-file-chip">Hoja: <strong>{sheet_name}</strong></span>
          <span class="pyc-file-chip {status_class}">Estado: <strong>{status_text}</strong></span>
        </div>
        """,
        unsafe_allow_html=True,
    )
    show_summary_chips(summary)


def _person_full_name(row: pd.Series) -> str:
    return f"{row['Nombre']} {row['Apellido']}".strip()


def _person_options(df: pd.DataFrame) -> list[str]:
    if df is None or df.empty:
        return []
    names = [_person_full_name(row) for _, row in df.iterrows()]
    return sorted(dict.fromkeys(name for name in names if name))


def build_supervisor_config(clean_df: pd.DataFrame) -> dict[str, dict[str, str]]:
    supervisors_by_state = {
        zone: _person_options(clean_df[clean_df["Estado"] == zone])
        for zone in ZONAS_CONFIGURADAS
    }

    config: dict[str, dict[str, str]] = {}
    for zone in ZONAS_CONFIGURADAS:
        detected_supervisors = supervisors_by_state.get(zone, [])
        config[zone] = {
            "supervisor": detected_supervisors[0] if detected_supervisors else "",
            "asistente": "",
        }

    return config


def show_generation_settings() -> dict[str, list[int]]:
    groups = sorted(build_turnos_por_grupo(GRUPO_PRIMER_SERVICIO_DEFAULT))
    selected_first_group = st.selectbox(
        "Grupo que inicia en el primer servicio",
        options=groups,
        index=groups.index(GRUPO_PRIMER_SERVICIO_DEFAULT) if GRUPO_PRIMER_SERVICIO_DEFAULT in groups else 0,
        key="grupo_primer_servicio",
    )
    turnos_por_grupo = build_turnos_por_grupo(selected_first_group)

    turnos_text = " | ".join(
        f"Grupo {grupo}: T{', T'.join(str(turno) for turno in turnos)}"
        for grupo, turnos in sorted(turnos_por_grupo.items())
    )
    st.caption(turnos_text)

    return turnos_por_grupo


def generate_and_store_schedule(
    available_df: pd.DataFrame,
    clean_df: pd.DataFrame,
    turnos_por_grupo: dict[str, list[int]],
) -> None:
    supervisor_config = build_supervisor_config(clean_df)
    schedule_df, alerts = generate_schedule(available_df, turnos_por_grupo)
    st.session_state["schedule_df"] = schedule_df
    st.session_state["alerts"] = alerts
    st.session_state["available_df"] = available_df
    st.session_state["people_df"] = clean_df
    st.session_state["turnos_por_grupo"] = turnos_por_grupo
    st.session_state["supervisor_config"] = supervisor_config


def show_alerts(alerts: list[str]) -> None:
    """Muestra alertas primero para que el coordinador atienda lo urgente."""
    alert_groups = _group_alerts(alerts)
    errors = alert_groups["Errores"]
    unassigned = alert_groups["Sin asignar"]
    warnings = alert_groups["Advertencias"]

    st.subheader("Validaciones y alertas")

    if not alerts:
        st.success("No se generaron alertas.")
        return

    tabs = st.tabs(
        [
            f"Sin asignar ({len(unassigned)})",
            f"Advertencias ({len(warnings)})",
            f"Errores ({len(errors)})",
        ]
    )

    with tabs[0]:
        _show_alert_group(unassigned, "warning")
    with tabs[1]:
        _show_alert_group(warnings, "info")
    with tabs[2]:
        _show_alert_group(errors, "error")


def _group_alerts(alerts: list[str]) -> dict[str, list[str]]:
    return {
        "Errores": [alert for alert in alerts if alert.startswith("ERROR")],
        "Sin asignar": [alert for alert in alerts if alert.startswith("SIN ASIGNAR")],
        "Advertencias": [
            alert
            for alert in alerts
            if not alert.startswith("ERROR") and not alert.startswith("SIN ASIGNAR")
        ],
    }


def _show_alert_group(alerts: list[str], level: str) -> None:
    if not alerts:
        st.success("Sin novedades en esta categoría.")
        return

    alert_df = pd.DataFrame({"Alerta": alerts})
    st.dataframe(alert_df, width="stretch", hide_index=True)


def show_alert_sidebar(alerts: list[str]) -> None:
    st.markdown("#### Alertas")
    alert_groups = _group_alerts(alerts)
    errors = alert_groups["Errores"]
    unassigned = alert_groups["Sin asignar"]
    warnings = alert_groups["Advertencias"]

    if not alerts:
        st.success("Sin alertas.")
        return

    tabs = st.tabs(
        [
            f"Sin asignar ({len(unassigned)})",
            f"Advertencias ({len(warnings)})",
            f"Errores ({len(errors)})",
        ]
    )

    with tabs[0]:
        _show_alert_cards(unassigned, "pyc-alert-warning")
    with tabs[1]:
        _show_alert_cards(warnings, "pyc-alert-info")
    with tabs[2]:
        _show_alert_cards(errors, "pyc-alert-danger")


def _show_alert_cards(alerts: list[str], css_class: str) -> None:
    if not alerts:
        st.success("Sin novedades.")
        return

    for alert in alerts[:10]:
        st.markdown(
            f'<div class="pyc-alert-card {css_class}">{alert}</div>',
            unsafe_allow_html=True,
        )

    if len(alerts) > 10:
        st.caption(f"+ {len(alerts) - 10} alertas más.")


def build_schedule_matrix(schedule_df: pd.DataFrame) -> pd.DataFrame:
    """Convierte la tabla plana en una matriz operativa por turnos."""
    rows = []
    positions_order = get_positions_df().reset_index().rename(columns={"index": "_orden"})
    turnos = sorted(TURNOS)

    for (zona, posicion, tipo), group in schedule_df.groupby(["Zona", "Posición", "Tipo"], sort=False):
        row = {
            "Zona": zona,
            "Posición": posicion,
            "Tipo": tipo,
        }

        for turno in turnos:
            turn_assignments = group[group["Turno"] == turno].sort_values("Slot")
            row[_turn_key(turno)] = "\n".join(_format_assignment(item) for _, item in turn_assignments.iterrows())

        rows.append(row)

    matrix_df = pd.DataFrame(rows)
    matrix_df = matrix_df.merge(
        positions_order[["zona", "posicion", "_orden"]],
        left_on=["Zona", "Posición"],
        right_on=["zona", "posicion"],
        how="left",
    )
    matrix_df = matrix_df.sort_values(["Zona", "_orden"]).drop(
        columns=["zona", "posicion", "_orden"]
    )
    return matrix_df


def show_schedule_matrix(
    schedule_df: pd.DataFrame,
    people_df: pd.DataFrame,
    available_df: pd.DataFrame,
    turnos_por_grupo: dict[str, list[int]] | None = None,
    supervisor_config: dict[str, dict[str, str]] | None = None,
) -> None:
    """Muestra una vista amable para coordinar posiciones por turno."""
    supervisor_config = supervisor_config or SUPERVISORES_ZONA
    zones = _ordered_zones(schedule_df)
    all_people = _person_options(people_df)

    _show_group_color_legend()
    tabs = st.tabs([_zone_tab_label(zone) for zone in zones])

    for tab, zone in zip(tabs, zones):
        with tab:
            supervisor, asistente = _show_zone_supervisor_controls(zone, supervisor_config, all_people)

            zone_schedule_df = schedule_df[schedule_df["Zona"] == zone]
            _show_editable_zone_matrix(zone, zone_schedule_df, available_df, turnos_por_grupo)


def _show_editable_zone_matrix(
    zone: str,
    zone_schedule_df: pd.DataFrame,
    available_df: pd.DataFrame,
    turnos_por_grupo: dict[str, list[int]] | None,
) -> None:
    edit_key = f"editing_zone_{zone}"
    if not st.session_state.get(edit_key, False):
        if st.button(f"Editar {zone}", key=f"edit_zone_{zone}", width="stretch"):
            st.session_state[edit_key] = True
            st.rerun()

        matrix_df = build_schedule_matrix(zone_schedule_df).drop(columns=["Zona"])
        st.dataframe(
            matrix_df.style.map(_style_matrix_cell),
            width="stretch",
            hide_index=True,
            column_config=_matrix_column_config(),
        )
        return

    volunteer_options, volunteer_lookup = _manual_assignment_options(available_df)
    editor_df = _editable_zone_matrix_df(zone_schedule_df)
    turn_keys = [_turn_key(turno) for turno in sorted(TURNOS)]
    row_key_columns = [f"_row_{turn_key}" for turn_key in turn_keys]

    edited_df = st.data_editor(
        editor_df,
        width="stretch",
        hide_index=True,
        disabled=["Posición", "Tipo"],
        column_config={
            **{row_key: None for row_key in row_key_columns},
            "Posición": st.column_config.TextColumn("Posición", width="medium"),
            "Tipo": st.column_config.TextColumn("Tipo", width="small"),
            **{
                turn_key: st.column_config.SelectboxColumn(
                    _turn_header(turno),
                    options=volunteer_options,
                    required=True,
                    width="medium",
                )
                for turno, turn_key in ((_turno, _turn_key(_turno)) for _turno in sorted(TURNOS))
            },
        },
        key=f"editable_zone_matrix_{zone}",
    )

    if st.button(f"Guardar cambios {zone}", key=f"save_zone_changes_{zone}", width="stretch"):
        updated_schedule, changed = _apply_zone_matrix_changes(
            st.session_state["schedule_df"],
            edited_df,
            volunteer_lookup,
            turnos_por_grupo,
        )
        if not changed:
            st.info("No hubo cambios para aplicar.")
            st.session_state[edit_key] = False
            st.rerun()
            return

        st.session_state["schedule_df"] = updated_schedule
        st.session_state["alerts"] = validate_schedule(updated_schedule, turnos_por_grupo)
        st.session_state[edit_key] = False
        st.success("Cambios aplicados. Alertas y programación actualizadas.")
        st.rerun()


def _editable_zone_matrix_df(zone_schedule_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    positions_order = get_positions_df().reset_index().rename(columns={"index": "_orden"})
    zone_schedule_df = zone_schedule_df.copy()
    zone_schedule_df["_schedule_index"] = zone_schedule_df.index
    zone_schedule_df = zone_schedule_df.merge(
        positions_order[["zona", "posicion", "_orden"]],
        left_on=["Zona", "Posición"],
        right_on=["zona", "posicion"],
        how="left",
    )
    multi_slot_positions = _multi_slot_positions(zone_schedule_df)

    grouped = zone_schedule_df.groupby(["Zona", "Posición", "Tipo", "Slot"], sort=False)
    for (_, position, row_type, slot), group in grouped:
        position_label = position
        if (position, row_type) in multi_slot_positions:
            position_label = f"{position} - Slot {int(slot)}"

        row = {
            "Posición": position_label,
            "Tipo": row_type,
            "_orden": group["_orden"].min(),
            "_slot": int(slot),
        }

        for turno in sorted(TURNOS):
            turn_key = _turn_key(turno)
            match = group[group["Turno"] == turno]
            if match.empty:
                row[turn_key] = FREE_NAME
                row[f"_row_{turn_key}"] = ""
                continue

            schedule_row = match.iloc[0]
            row[turn_key] = _manual_assignment_label(schedule_row)
            row[f"_row_{turn_key}"] = str(schedule_row["_schedule_index"])

        rows.append(row)

    matrix_df = pd.DataFrame(rows).sort_values(["_orden", "_slot"]).drop(columns=["_orden", "_slot"])
    return matrix_df


def _multi_slot_positions(zone_schedule_df: pd.DataFrame) -> set[tuple[str, str]]:
    slot_counts = zone_schedule_df.groupby(["Posición", "Tipo"])["Slot"].nunique()
    return {key for key, count in slot_counts.items() if count > 1}


def _apply_zone_matrix_changes(
    schedule_df: pd.DataFrame,
    edited_df: pd.DataFrame,
    volunteer_lookup: dict[str, pd.Series | str],
    turnos_por_grupo: dict[str, list[int]] | None,
) -> tuple[pd.DataFrame, bool]:
    updated_df = schedule_df.copy()
    changed = False

    for _, row in edited_df.iterrows():
        for turno in sorted(TURNOS):
            turn_key = _turn_key(turno)
            row_index = row.get(f"_row_{turn_key}", "")
            if row_index == "":
                continue

            row_index = int(row_index)
            selected_label = row[turn_key]
            current_label = _manual_assignment_label(updated_df.loc[row_index])
            if selected_label == current_label:
                continue

            updated_df = _apply_assignment_label(
                updated_df,
                row_index,
                selected_label,
                volunteer_lookup,
                turnos_por_grupo=turnos_por_grupo,
            )
            changed = True

    return updated_df, changed


def _apply_assignment_label(
    schedule_df: pd.DataFrame,
    row_index: int,
    selected_label: str,
    volunteer_lookup: dict[str, pd.Series | str],
    turnos_por_grupo: dict[str, list[int]] | None,
) -> pd.DataFrame:
    selected = volunteer_lookup[selected_label]
    updated_df = schedule_df.copy()

    if isinstance(selected, str) and selected == UNASSIGNED_NAME:
        updated_df.loc[row_index, ["Grupo", "Nombre", "Apellido", "Género", "Estado", "Observaciones"]] = [
            "",
            UNASSIGNED_NAME,
            "",
            "",
            "",
            "",
        ]
        return updated_df

    if isinstance(selected, str) and selected == FREE_NAME:
        updated_df.loc[row_index, ["Grupo", "Nombre", "Apellido", "Género", "Estado", "Observaciones"]] = [
            "",
            FREE_NAME,
            "",
            "",
            "",
            "",
        ]
        return updated_df

    updated_df, _ = apply_manual_assignment(updated_df, row_index, selected, turnos_por_grupo)
    return updated_df


def _show_group_color_legend() -> None:
    st.markdown(
        """
        <div class="pyc-volunteer-strip">
          <span class="pyc-volunteer-chip pyc-group-chip-a">Azul <strong>Grupo A</strong></span>
          <span class="pyc-volunteer-chip pyc-group-chip-b">Verde <strong>Grupo B</strong></span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _show_zone_supervisor_controls(
    zone: str,
    supervisor_config: dict[str, dict[str, str]],
    all_people: list[str],
) -> tuple[str, str]:
    config = supervisor_config.setdefault(zone, {"supervisor": "", "asistente": ""})
    supervisor_value = st.session_state.get(f"zone_supervisor_{zone}", config.get("supervisor", ""))
    assistant_value = st.session_state.get(f"zone_assistant_{zone}", config.get("asistente", ""))
    supervisor_value = "" if supervisor_value == "sin asignar" else supervisor_value
    assistant_value = "" if assistant_value == "sin asignar" else assistant_value

    supervisor_options = ["sin asignar", *all_people]
    supervisor_index = _option_index(supervisor_options, supervisor_value)

    selected_supervisors = {
        zone_config.get("supervisor", "")
        for zone_config in supervisor_config.values()
        if zone_config.get("supervisor")
    }
    assistant_people = [
        person
        for person in all_people
        if person != supervisor_value and person not in selected_supervisors
    ]
    assistant_options = ["sin asignar", *assistant_people]
    assistant_index = _option_index(assistant_options, assistant_value)

    cols = st.columns([1.0, 1.0], gap="small")
    with cols[0]:
        supervisor = st.selectbox(
            f"**Supervisor {zone}**",
            options=supervisor_options,
            index=supervisor_index,
            key=f"zone_supervisor_{zone}",
        )
    with cols[1]:
        asistente = st.selectbox(
            f"**Asistente {zone}**",
            options=assistant_options,
            index=assistant_index,
            key=f"zone_assistant_{zone}",
        )

    config["supervisor"] = "" if supervisor == "sin asignar" else supervisor
    config["asistente"] = "" if asistente == "sin asignar" else asistente
    st.session_state["supervisor_config"] = supervisor_config
    return config["supervisor"], config["asistente"]


def _option_index(options: list[str], value: str) -> int:
    if value and value in options:
        return options.index(value)
    return 0


def _display_person_or_unassigned(value: str) -> str:
    return value if value else "sin asignar"


def show_assignment_coverage(schedule_df: pd.DataFrame, available_df: pd.DataFrame) -> None:
    """Muestra que voluntarios disponibles quedaron sin uso en la programacion."""
    assigned_people = schedule_df[~schedule_df["Nombre"].isin([UNASSIGNED_NAME, FREE_NAME])][
        ["Grupo", "Nombre", "Apellido"]
    ].drop_duplicates()

    available_keys = available_df.copy()
    assigned_keys = assigned_people.copy()

    for df in [available_keys, assigned_keys]:
        df["_persona_key"] = (
            df["Grupo"].astype(str).str.strip()
            + "|"
            + df["Nombre"].astype(str).str.strip()
            + "|"
            + df["Apellido"].astype(str).str.strip()
        )

    unused_df = available_keys[
        ~available_keys["_persona_key"].isin(assigned_keys["_persona_key"])
    ].drop(columns=["_persona_key"])

    summary_left, summary_right = st.columns([2.2, 1.0], gap="small")
    with summary_left:
        st.markdown(
            f"""
            <div class="pyc-volunteer-strip">
              <span class="pyc-volunteer-chip">Disponibles <strong>{len(available_df)}</strong></span>
              <span class="pyc-volunteer-chip">Asignados <strong>{len(assigned_people)}</strong></span>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with summary_right:
        with st.popover(f"Sin asignación: {len(unused_df)}", width="stretch"):
            if unused_df.empty:
                st.success("Todos los voluntarios disponibles fueron asignados al menos una vez.")
            else:
                st.dataframe(unused_df, width="stretch", hide_index=True)


def _format_assignment(row: pd.Series) -> str:
    if row["Nombre"] == UNASSIGNED_NAME:
        return UNASSIGNED_NAME
    if row["Nombre"] == FREE_NAME:
        return FREE_NAME

    full_name = f"{row['Nombre']} {row['Apellido']}".strip()
    tags = [row["Grupo"]]

    if row["Estado"] in set(ZONAS_CONFIGURADAS) | {"apoyo"}:
        tags.append(row["Estado"])

    return f"{full_name} ({', '.join(tags)})"


def _turn_key(turno: int) -> str:
    return f"T{int(turno)}"


def _turn_header(turno: int) -> str:
    return f"T{int(turno)}: {TURNOS[int(turno)].replace('-', ' - ')}"


def _ordered_zones(schedule_df: pd.DataFrame) -> list[str]:
    configured_order = list(ZONAS_CONFIGURADAS)
    present_zones = list(dict.fromkeys(schedule_df["Zona"].astype(str)))
    ordered = [zone for zone in configured_order if zone in present_zones]
    ordered.extend(zone for zone in present_zones if zone not in ordered)
    return ordered


def _zone_tab_label(zone: str) -> str:
    title = TITULOS_ZONA.get(zone, zone)
    if f"({zone})" in title:
        suffix = title.split(f"({zone})", 1)[1].strip()
        return f"{zone} {suffix}".strip()
    return title


def _matrix_column_config() -> dict:
    config = {
        "Posición": st.column_config.TextColumn("Posición", width="medium"),
        "Tipo": st.column_config.TextColumn("Tipo", width="small"),
    }
    for turno in sorted(TURNOS):
        config[_turn_key(turno)] = st.column_config.TextColumn(_turn_header(turno), width="medium")
    return config


def _style_matrix_cell(value: object) -> str:
    text = str(value)

    if UNASSIGNED_NAME in text:
        return "background-color: #fde2e1; color: #9f1239; font-weight: 700;"

    if FREE_NAME in text:
        return "background-color: #f1f3f5; color: #6c757d; font-style: italic;"

    if text == "critica":
        return "background-color: #e5f0dc; color: #31572c; font-weight: 700;"

    if text == "refuerzo":
        return "background-color: #fff3cd; color: #8a5a00; font-weight: 700;"

    has_group_a = "(A)" in text or "(A," in text
    has_group_b = "(B)" in text or "(B," in text
    if has_group_a and not has_group_b:
        return "background-color: #dbeafe; color: #1e3a8a; font-weight: 700;"
    if has_group_b and not has_group_a:
        return "background-color: #dcfce7; color: #14532d; font-weight: 700;"
    if has_group_a and has_group_b:
        return "background-color: #ede9fe; color: #4c1d95; font-weight: 700;"

    if text.strip() and text not in {"Zona", "Posición", "Tipo"}:
        return "background-color: #d7e8f7; color: #1f3b63; font-weight: 650;"

    return ""


def show_manual_reassignment(
    schedule_df: pd.DataFrame,
    available_df: pd.DataFrame,
    turnos_por_grupo: dict[str, list[int]] | None = None,
) -> None:
    st.subheader("Reasignación manual")

    with st.expander("Editar asignaciones", expanded=False):
        volunteer_options, volunteer_lookup = _manual_assignment_options(available_df)
        editor_df = _manual_editor_df(schedule_df)
        editor_df["Asignacion manual"] = editor_df["_row_index"].apply(
            lambda row_index: _manual_assignment_label(schedule_df.loc[row_index])
        )

        edited_df = st.data_editor(
            editor_df,
            width="stretch",
            hide_index=True,
            disabled=["_row_index", "Zona", "Turno", "Horario", "Posición", "Slot", "Tipo", "Obligatorio"],
            column_config={
                "_row_index": None,
                "Asignacion manual": st.column_config.SelectboxColumn(
                    "Asignacion manual",
                    options=volunteer_options,
                    required=True,
                    width="large",
                ),
            },
            key="manual_assignment_editor",
        )

        if st.button("Aplicar cambios manuales", width="stretch"):
            updated_schedule, changed = _apply_manual_editor_changes(
                schedule_df,
                edited_df,
                volunteer_lookup,
                turnos_por_grupo,
            )
            if not changed:
                st.info("No hubo cambios para aplicar.")
                return

            st.session_state["schedule_df"] = updated_schedule
            st.session_state["alerts"] = validate_schedule(updated_schedule, turnos_por_grupo)
            st.success("Cambios manuales aplicados. Alertas y programación actualizadas.")
            st.rerun()


def _manual_editor_df(schedule_df: pd.DataFrame) -> pd.DataFrame:
    columns = ["Zona", "Turno", "Horario", "Posición", "Slot", "Tipo", "Obligatorio"]
    editor_df = schedule_df.reset_index().rename(columns={"index": "_row_index"})
    return editor_df[["_row_index", *columns]].sort_values(["Turno", "Zona", "Posición", "Slot"])


def _manual_assignment_options(available_df: pd.DataFrame) -> tuple[list[str], dict[str, pd.Series | str]]:
    options = [UNASSIGNED_NAME, FREE_NAME]
    lookup: dict[str, pd.Series | str] = {
        UNASSIGNED_NAME: UNASSIGNED_NAME,
        FREE_NAME: FREE_NAME,
    }

    for _, row in available_df.sort_values(["Grupo", "Nombre", "Apellido"]).iterrows():
        label = _manual_assignment_label(row)
        options.append(label)
        lookup[label] = row

    return options, lookup


def _manual_assignment_label(row: pd.Series) -> str:
    if row["Nombre"] in {UNASSIGNED_NAME, FREE_NAME}:
        return row["Nombre"]
    return (
        f"{row['Nombre']} {row['Apellido']} "
        f"({row['Grupo']} / {row['Género']} / {row['Estado']})"
    )


def _apply_manual_editor_changes(
    schedule_df: pd.DataFrame,
    edited_df: pd.DataFrame,
    volunteer_lookup: dict[str, pd.Series | str],
    turnos_por_grupo: dict[str, list[int]] | None = None,
) -> tuple[pd.DataFrame, bool]:
    updated_df = schedule_df.copy()
    changed = False

    for _, row in edited_df.iterrows():
        row_index = int(row["_row_index"])
        selected_label = row["Asignacion manual"]
        current_label = _manual_assignment_label(updated_df.loc[row_index])
        if selected_label == current_label:
            continue

        selected = volunteer_lookup[selected_label]
        if isinstance(selected, str) and selected == UNASSIGNED_NAME:
            updated_df.loc[row_index, ["Grupo", "Nombre", "Apellido", "Género", "Estado", "Observaciones"]] = [
                "",
                UNASSIGNED_NAME,
                "",
                "",
                "",
                "",
            ]
        elif isinstance(selected, str) and selected == FREE_NAME:
            updated_df.loc[row_index, ["Grupo", "Nombre", "Apellido", "Género", "Estado", "Observaciones"]] = [
                "",
                FREE_NAME,
                "",
                "",
                "",
                "",
            ]
        else:
            updated_df, _ = apply_manual_assignment(updated_df, row_index, selected, turnos_por_grupo)
        changed = True

    return updated_df, changed


def _slot_label(index: int, row: pd.Series) -> str:
    assignment = _format_assignment(row)
    return (
        f"Fila {index} | T{row['Turno']} {row['Horario']} | {row['Zona']} | "
        f"{row['Posición']} | Slot {row['Slot']} | {assignment}"
    )


def show_schedule_results(
    schedule_df: pd.DataFrame,
    alerts: list[str],
    available_df: pd.DataFrame,
    people_df: pd.DataFrame,
    turnos_por_grupo: dict[str, list[int]] | None = None,
    supervisor_config: dict[str, dict[str, str]] | None = None,
) -> None:
    center_col, right_col = st.columns([3.1, 0.9], gap="small")

    with center_col:
        supervisor_config = _supervisor_config_from_state(supervisor_config or SUPERVISORES_ZONA)
        excel_bytes = build_output_excel(schedule_df, alerts, available_df, supervisor_config)

        title_col, download_col = st.columns([2.6, 0.7], gap="small")
        with title_col:
            st.subheader("Programación por posiciones")
        with download_col:
            st.download_button(
                "Descargar Excel",
                data=excel_bytes,
                file_name=_download_file_name(),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )

        show_assignment_coverage(schedule_df, available_df)
        show_schedule_matrix(schedule_df, people_df, available_df, turnos_por_grupo, supervisor_config)

        with st.expander("Programación detallada", expanded=False):
            st.dataframe(schedule_df, width="stretch")

    with right_col:
        show_alert_sidebar(alerts)


def _supervisor_config_from_state(
    supervisor_config: dict[str, dict[str, str]],
) -> dict[str, dict[str, str]]:
    config = {
        zone: values.copy()
        for zone, values in supervisor_config.items()
    }
    for zone in ZONAS_CONFIGURADAS:
        zone_config = config.setdefault(zone, {"supervisor": "", "asistente": ""})
        supervisor = st.session_state.get(f"zone_supervisor_{zone}", zone_config.get("supervisor", ""))
        asistente = st.session_state.get(f"zone_assistant_{zone}", zone_config.get("asistente", ""))
        zone_config["supervisor"] = "" if supervisor == "sin asignar" else supervisor
        zone_config["asistente"] = "" if asistente == "sin asignar" else asistente
    st.session_state["supervisor_config"] = config
    return config


def _download_file_name() -> str:
    return f"programacion_pyc_{datetime.now():%Y_%m}.xlsx"


def main() -> None:
    st.set_page_config(page_title="PYC Organizador de Servicio", layout="wide")
    inject_app_styles()

    has_uploaded_file = "uploaded_file_bytes" in st.session_state
    uploaded_file_name = st.session_state.get("uploaded_file_name", "")
    uploaded_sheet_name = st.session_state.get("uploaded_sheet_name", "")
    read_error_title = ""
    read_error_detail = ""

    if has_uploaded_file:
        try:
            raw_df = read_volunteers(BytesIO(st.session_state["uploaded_file_bytes"]))
            uploaded_sheet_name = raw_df.attrs.get("sheet_name", uploaded_sheet_name)
            st.session_state["uploaded_sheet_name"] = uploaded_sheet_name
        except ValueError as exc:
            raw_df = None
            read_error_title = "No se pudo leer la hoja de voluntarios."
            read_error_detail = str(exc)
        except Exception as exc:
            raw_df = None
            read_error_title = "No se pudo leer el archivo Excel."
            read_error_detail = str(exc)

        validation_errors = [read_error_detail] if read_error_detail else validate_volunteers(raw_df)
        if not validation_errors:
            clean_df = clean_volunteers(raw_df)
            available_df = get_available_volunteers(clean_df)
            unavailable_df = get_unavailable_volunteers(clean_df)
            summary = get_summary(clean_df, available_df)
        else:
            clean_df = available_df = unavailable_df = summary = None
    else:
        raw_df = clean_df = available_df = unavailable_df = summary = validation_errors = None

    max_step = 1
    if has_uploaded_file and not validation_errors:
        max_step = 2
    if "schedule_df" in st.session_state:
        max_step = 3

    active_step = get_requested_step(max_step)
    show_top_nav(active_step, max_step, has_step_error=active_step == 1 and bool(validation_errors))

    if active_step == 1:
        st.markdown("#### Resumen de archivo")
        if has_uploaded_file:
            show_loaded_file_badge(type("UploadedFileName", (), {"name": uploaded_file_name})())
            st.markdown('<div class="pyc-file-uploader">', unsafe_allow_html=True)
            replacement_file = st.file_uploader(
                "Cambiar archivo",
                type=["xlsx"],
                help="Carga otro Excel para reemplazar el archivo actual.",
                key="replacement_file",
            )
            st.markdown("</div>", unsafe_allow_html=True)
            if replacement_file is not None:
                handle_uploaded_file(replacement_file, target_step=2)

            if read_error_title:
                st.error(read_error_title)
                st.caption(read_error_detail)
            elif validation_errors:
                st.markdown("#### Validación de archivo")
                st.error("El archivo tiene errores de validación.")
                st.dataframe(pd.DataFrame({"Error": validation_errors}), width="stretch", hide_index=True)
            else:
                st.markdown("#### Validación de archivo")
                st.success("Validación completada. El archivo tiene la estructura esperada.")
                if st.button("Continuar a generar programación", type="primary", width="stretch"):
                    st.query_params["step"] = "2"
                    st.rerun()
        else:
            st.info("Carga un archivo Excel para comenzar.")
            st.markdown('<div class="pyc-file-uploader">', unsafe_allow_html=True)
            uploaded_file = st.file_uploader(
                "Archivo de voluntarios",
                type=["xlsx"],
                help="Si el archivo tiene una sola hoja, se usará esa hoja. Si tiene varias, debe existir una hoja llamada Voluntarios.",
            )
            st.markdown("</div>", unsafe_allow_html=True)
            if uploaded_file is not None:
                handle_uploaded_file(uploaded_file, target_step=2)
        return

    if active_step == 3 and "schedule_df" in st.session_state:
        show_schedule_results(
            st.session_state["schedule_df"],
            st.session_state["alerts"],
            st.session_state["available_df"],
            st.session_state.get("people_df", st.session_state["available_df"]),
            st.session_state.get("turnos_por_grupo"),
            st.session_state.get("supervisor_config"),
        )
        return

    st.markdown("#### Generar programación")
    show_file_summary(
        type("UploadedFileName", (), {"name": uploaded_file_name})(),
        summary,
        uploaded_sheet_name or "Voluntarios",
    )
    turnos_por_grupo = show_generation_settings()
    action_col, _ = st.columns([0.22, 0.78])
    with action_col:
        with st.form("generate_schedule_form"):
            generate_clicked = st.form_submit_button("Generar", type="primary", width="stretch")
    if generate_clicked:
        generate_and_store_schedule(available_df, clean_df, turnos_por_grupo)
        st.query_params["step"] = "3"
        st.rerun()

    with st.expander("Voluntarios disponibles", expanded=False):
        st.dataframe(available_df, width="stretch")

    with st.expander("Voluntarios no disponibles", expanded=False):
        st.dataframe(unavailable_df, width="stretch")


if __name__ == "__main__":
    main()
